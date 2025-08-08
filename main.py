#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excelレコード検索・抽出ツール（標準ライブラリ版）

【機能】
- Excelファイルから名称で検索
- 完全一致・部分一致・曖昧検索に対応
- 検索結果をプレビュー表示
- 結果をExcelファイルに保存

【必要なライブラリ】
pip install pandas openpyxl
（標準ライブラリ + pandas/openpyxl のみで動作）

作成者: [あなたの名前]
更新日: 2024-08-08
目的: 大量データから効率的にレコード抽出
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
import difflib  # 標準ライブラリの文字列類似度計算
import openpyxl
from openpyxl.styles import Font, PatternFill

class ExcelSearchTool:
    """Excelファイル検索・抽出ツール（標準ライブラリ版）"""
    
    def __init__(self, root):
        self.root = root
        self.root.title("Excel レコード検索・抽出ツール（標準版）")
        self.root.geometry("1000x800")
        
        # データ保存用変数
        self.df = None  # 読み込んだExcelデータ
        self.search_results = None  # 検索結果
        self.columns = []  # 列名リスト
        
        # GUI変数
        self.input_file_path = tk.StringVar()
        self.output_file_path = tk.StringVar()
        self.search_column = tk.StringVar()
        self.search_keyword = tk.StringVar()
        self.search_mode = tk.StringVar(value="exact")
        self.similarity_threshold = tk.IntVar(value=80)
        self.ignore_case = tk.BooleanVar(value=True)  # 大文字小文字を無視
        
        self.setup_ui()
    
    def setup_ui(self):
        """UIセットアップ"""
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # ファイル選択エリア
        self.create_file_area(main_frame)
        
        # 検索設定エリア
        self.create_search_area(main_frame)
        
        # プレビューエリア
        self.create_preview_area(main_frame)
        
        # ボタンエリア
        self.create_button_area(main_frame)
        
        # ステータスエリア
        self.create_status_area(main_frame)
        
        # グリッド設定
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
    
    def create_file_area(self, parent):
        """ファイル選択エリア作成"""
        file_frame = ttk.LabelFrame(parent, text="ファイル選択", padding="5")
        file_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)
        
        # 入力ファイル
        ttk.Label(file_frame, text="検索対象Excelファイル:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        ttk.Entry(file_frame, textvariable=self.input_file_path, width=60).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 5))
        ttk.Button(file_frame, text="参照", command=self.browse_input_file).grid(row=0, column=2)
        
        # 出力ファイル
        ttk.Label(file_frame, text="結果保存先:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(5, 0))
        ttk.Entry(file_frame, textvariable=self.output_file_path, width=60).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(0, 5), pady=(5, 0))
        ttk.Button(file_frame, text="参照", command=self.browse_output_file).grid(row=1, column=2, pady=(5, 0))
    
    def create_search_area(self, parent):
        """検索設定エリア作成"""
        search_frame = ttk.LabelFrame(parent, text="検索設定", padding="5")
        search_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 10))
        search_frame.columnconfigure(1, weight=1)
        
        # 検索対象列
        ttk.Label(search_frame, text="検索対象列:").grid(row=0, column=0, sticky=tk.W, padx=(0, 5))
        self.column_combo = ttk.Combobox(search_frame, textvariable=self.search_column, width=20)
        self.column_combo.grid(row=0, column=1, sticky=tk.W, padx=(0, 10))
        
        # 検索キーワード
        ttk.Label(search_frame, text="検索キーワード:").grid(row=0, column=2, sticky=tk.W, padx=(10, 5))
        ttk.Entry(search_frame, textvariable=self.search_keyword, width=30).grid(row=0, column=3, sticky=(tk.W, tk.E))
        
        # 検索モード
        ttk.Label(search_frame, text="検索モード:").grid(row=1, column=0, sticky=tk.W, padx=(0, 5), pady=(5, 0))
        mode_frame = ttk.Frame(search_frame)
        mode_frame.grid(row=1, column=1, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # ラジオボタンで検索モード選択
        ttk.Radiobutton(mode_frame, text="完全一致", variable=self.search_mode, value="exact").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(mode_frame, text="部分一致", variable=self.search_mode, value="partial").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Radiobutton(mode_frame, text="曖昧検索", variable=self.search_mode, value="fuzzy").pack(side=tk.LEFT, padx=(0, 10))
        
        # 検索オプション
        ttk.Checkbutton(mode_frame, text="大文字小文字を無視", variable=self.ignore_case).pack(side=tk.LEFT, padx=(20, 10))
        
        # 曖昧検索の類似度設定
        ttk.Label(mode_frame, text="類似度(%)：").pack(side=tk.LEFT, padx=(10, 5))
        similarity_scale = ttk.Scale(mode_frame, from_=50, to=100, variable=self.similarity_threshold, orient=tk.HORIZONTAL, length=100)
        similarity_scale.pack(side=tk.LEFT, padx=(0, 5))
        self.similarity_label = ttk.Label(mode_frame, text="80%")
        self.similarity_label.pack(side=tk.LEFT)
        
        # スケールの値変更時のイベント
        similarity_scale.configure(command=self.update_similarity_label)
    
    def create_preview_area(self, parent):
        """プレビューエリア作成"""
        preview_frame = ttk.LabelFrame(parent, text="データプレビュー・検索結果", padding="5")
        preview_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(1, weight=1)
        
        # タブでデータと検索結果を分離
        self.notebook = ttk.Notebook(preview_frame)
        self.notebook.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # データプレビュータブ
        self.data_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.data_frame, text="元データ")
        
        self.data_text = scrolledtext.ScrolledText(self.data_frame, height=20, width=100)
        self.data_text.pack(fill=tk.BOTH, expand=True)
        
        # 検索結果タブ
        self.result_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.result_frame, text="検索結果")
        
        self.result_text = scrolledtext.ScrolledText(self.result_frame, height=20, width=100)
        self.result_text.pack(fill=tk.BOTH, expand=True)
    
    def create_button_area(self, parent):
        """ボタンエリア作成"""
        button_frame = ttk.Frame(parent)
        button_frame.grid(row=3, column=0, columnspan=2, pady=(0, 10))
        
        ttk.Button(button_frame, text="ファイル読込", command=self.load_excel_file).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="検索実行", command=self.execute_search, style="Accent.TButton").pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="結果保存", command=self.save_results).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="クリア", command=self.clear_all).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(button_frame, text="サンプル作成", command=self.create_sample).pack(side=tk.LEFT)
    
    def create_status_area(self, parent):
        """ステータスエリア作成"""
        self.status_var = tk.StringVar(value="ファイルを選択してください")
        status_frame = ttk.Frame(parent)
        status_frame.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E))
        status_frame.columnconfigure(0, weight=1)
        
        ttk.Label(status_frame, textvariable=self.status_var, relief=tk.SUNKEN).grid(row=0, column=0, sticky=(tk.W, tk.E))
    
    def browse_input_file(self):
        """入力ファイル選択"""
        filename = filedialog.askopenfilename(
            title="検索対象Excelファイルを選択",
            filetypes=[("Excelファイル", "*.xlsx *.xls"), ("すべてのファイル", "*.*")]
        )
        if filename:
            self.input_file_path.set(filename)
            # 出力ファイル名を自動生成
            if not self.output_file_path.get():
                base_name = Path(filename).stem
                output_path = Path(filename).parent / f"{base_name}_検索結果.xlsx"
                self.output_file_path.set(str(output_path))
    
    def browse_output_file(self):
        """出力ファイル選択"""
        filename = filedialog.asksaveasfilename(
            title="結果保存先を選択",
            defaultextension=".xlsx",
            filetypes=[("Excelファイル", "*.xlsx"), ("すべてのファイル", "*.*")]
        )
        if filename:
            self.output_file_path.set(filename)
    
    def update_similarity_label(self, value):
        """類似度ラベル更新"""
        self.similarity_label.config(text=f"{int(float(value))}%")
    
    def load_excel_file(self):
        """Excelファイル読み込み"""
        if not self.input_file_path.get():
            messagebox.showwarning("警告", "Excelファイルを選択してください。")
            return
        
        try:
            self.status_var.set("ファイル読み込み中...")
            self.root.update()
            
            # Excelファイル読み込み
            file_path = self.input_file_path.get()
            if file_path.endswith('.xlsx'):
                self.df = pd.read_excel(file_path, engine='openpyxl')
            else:
                self.df = pd.read_excel(file_path, engine='xlrd')
            
            # 列名を取得してコンボボックスに設定
            self.columns = list(self.df.columns)
            self.column_combo['values'] = self.columns
            if self.columns:
                self.search_column.set(self.columns[0])  # 最初の列を選択
            
            # データプレビュー表示
            self.display_data_preview()
            
            self.status_var.set(f"ファイル読み込み完了: {len(self.df)}行 × {len(self.df.columns)}列")
            
        except Exception as e:
            messagebox.showerror("エラー", f"ファイル読み込みエラー: {e}")
            self.status_var.set("ファイル読み込みエラー")
    
    def display_data_preview(self):
        """データプレビュー表示"""
        self.data_text.delete(1.0, tk.END)
        
        if self.df is None:
            self.data_text.insert(tk.END, "データが読み込まれていません。")
            return
        
        # ヘッダー表示
        headers = " | ".join([f"{col:<15}" for col in self.df.columns])
        self.data_text.insert(tk.END, f"列名: {headers}\n")
        self.data_text.insert(tk.END, "-" * len(headers) + "\n")
        
        # 最初の20行を表示
        for idx, row in self.df.head(20).iterrows():
            row_data = " | ".join([f"{str(val):<15}" for val in row.values])
            self.data_text.insert(tk.END, f"[{idx+1:3d}] {row_data}\n")
        
        if len(self.df) > 20:
            self.data_text.insert(tk.END, f"\n... 他 {len(self.df) - 20} 行")
    
    def execute_search(self):
        """検索実行"""
        if self.df is None:
            messagebox.showwarning("警告", "まずExcelファイルを読み込んでください。")
            return
        
        if not self.search_keyword.get().strip():
            messagebox.showwarning("警告", "検索キーワードを入力してください。")
            return
        
        if not self.search_column.get():
            messagebox.showwarning("警告", "検索対象列を選択してください。")
            return
        
        try:
            self.status_var.set("検索中...")
            self.root.update()
            
            keyword = self.search_keyword.get().strip()
            column = self.search_column.get()
            mode = self.search_mode.get()
            
            # 検索実行
            if mode == "exact":
                self.search_results = self.exact_search(keyword, column)
            elif mode == "partial":
                self.search_results = self.partial_search(keyword, column)
            elif mode == "fuzzy":
                threshold = self.similarity_threshold.get()
                self.search_results = self.fuzzy_search(keyword, column, threshold)
            
            # 結果表示
            self.display_search_results()
            
            # 結果タブに切り替え
            self.notebook.select(1)
            
            result_count = len(self.search_results) if self.search_results is not None else 0
            self.status_var.set(f"検索完了: {result_count}件の結果が見つかりました")
            
        except Exception as e:
            messagebox.showerror("エラー", f"検索エラー: {e}")
            self.status_var.set("検索エラー")
    
    def exact_search(self, keyword: str, column: str) -> pd.DataFrame:
        """完全一致検索"""
        if self.ignore_case.get():
            # 大文字小文字を無視する場合
            mask = self.df[column].astype(str).str.lower() == keyword.lower()
        else:
            # 大文字小文字を区別する場合
            mask = self.df[column].astype(str) == keyword
        return self.df[mask].copy()
    
    def partial_search(self, keyword: str, column: str) -> pd.DataFrame:
        """部分一致検索"""
        if self.ignore_case.get():
            # 大文字小文字を無視する場合
            mask = self.df[column].astype(str).str.lower().str.contains(keyword.lower(), na=False)
        else:
            # 大文字小文字を区別する場合
            mask = self.df[column].astype(str).str.contains(keyword, na=False)
        return self.df[mask].copy()
    
    def fuzzy_search(self, keyword: str, column: str, threshold: int) -> pd.DataFrame:
        """曖昧検索（標準ライブラリ版）"""
        def calculate_similarity(text):
            if pd.isna(text):
                return 0
            
            # 大文字小文字の処理
            if self.ignore_case.get():
                text_processed = str(text).lower()
                keyword_processed = keyword.lower()
            else:
                text_processed = str(text)
                keyword_processed = keyword
            
            # difflibを使用した類似度計算
            similarity = difflib.SequenceMatcher(None, keyword_processed, text_processed).ratio()
            return int(similarity * 100)
        
        # 類似度を計算
        similarities = []
        matching_indices = []
        
        for idx, row in self.df.iterrows():
            similarity = calculate_similarity(row[column])
            if similarity >= threshold:
                similarities.append(similarity)
                matching_indices.append(idx)
        
        if not matching_indices:
            return pd.DataFrame()
        
        # 結果を取得してソート
        results = self.df.loc[matching_indices].copy()
        results['類似度'] = similarities
        results = results.sort_values('類似度', ascending=False)
        
        return results
    
    def advanced_fuzzy_search(self, keyword: str, text: str) -> int:
        """高度な曖昧検索アルゴリズム（標準ライブラリのみ）"""
        if not text or not keyword:
            return 0
        
        # 大文字小文字の処理
        if self.ignore_case.get():
            text = text.lower()
            keyword = keyword.lower()
        
        # 完全一致
        if keyword == text:
            return 100
        
        # 部分一致
        if keyword in text or text in keyword:
            return 85
        
        # 文字レベルの類似度（SequenceMatcher）
        seq_ratio = difflib.SequenceMatcher(None, keyword, text).ratio()
        
        # 単語レベルの類似度
        keyword_words = set(keyword.split())
        text_words = set(text.split())
        if keyword_words and text_words:
            word_ratio = len(keyword_words & text_words) / len(keyword_words | text_words)
        else:
            word_ratio = 0
        
        # 総合スコア
        final_score = max(seq_ratio * 100, word_ratio * 100)
        return int(final_score)
    
    def display_search_results(self):
        """検索結果表示"""
        self.result_text.delete(1.0, tk.END)
        
        if self.search_results is None or len(self.search_results) == 0:
            self.result_text.insert(tk.END, "検索結果が見つかりませんでした。\n\n")
            self.result_text.insert(tk.END, "検索のヒント:\n")
            self.result_text.insert(tk.END, "• 完全一致: 文字列が完全に一致する場合のみ\n")
            self.result_text.insert(tk.END, "• 部分一致: キーワードが含まれている場合\n")
            self.result_text.insert(tk.END, "• 曖昧検索: 似たような文字列も含める（誤字脱字に強い）\n")
            self.result_text.insert(tk.END, "• 大文字小文字を無視: チェックで区別しない\n")
            return
        
        # ヘッダー表示
        headers = " | ".join([f"{col:<15}" for col in self.search_results.columns])
        self.result_text.insert(tk.END, f"検索結果: {len(self.search_results)}件\n")
        self.result_text.insert(tk.END, f"検索モード: {self.search_mode.get()}\n")
        self.result_text.insert(tk.END, f"列名: {headers}\n")
        self.result_text.insert(tk.END, "-" * len(headers) + "\n")
        
        # 結果表示（最大50行）
        for idx, row in self.search_results.head(50).iterrows():
            row_data = " | ".join([f"{str(val):<15}" for val in row.values])
            self.result_text.insert(tk.END, f"[{idx+1:3d}] {row_data}\n")
        
        if len(self.search_results) > 50:
            self.result_text.insert(tk.END, f"\n... 他 {len(self.search_results) - 50} 行（保存時には全件出力されます）")
    
    def save_results(self):
        """検索結果をExcelファイルに保存"""
        if self.search_results is None or len(self.search_results) == 0:
            messagebox.showwarning("警告", "保存する検索結果がありません。")
            return
        
        if not self.output_file_path.get():
            messagebox.showwarning("警告", "保存先ファイルを指定してください。")
            return
        
        try:
            self.status_var.set("ファイル保存中...")
            self.root.update()
            
            output_path = self.output_file_path.get()
            
            # Excelファイルに保存（スタイル付き）
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 検索結果を保存
                self.search_results.to_excel(writer, sheet_name='検索結果', index=False)
                
                # 検索条件を保存
                search_info = pd.DataFrame({
                    '検索条件': ['検索キーワード', '検索対象列', '検索モード', '大文字小文字無視'],
                    '値': [self.search_keyword.get(), self.search_column.get(), 
                          self.search_mode.get(), str(self.ignore_case.get())]
                })
                search_info.to_excel(writer, sheet_name='検索条件', index=False)
                
                # 元データも参考用に保存
                if self.df is not None:
                    self.df.head(1000).to_excel(writer, sheet_name='元データ（抜粋）', index=False)
                
                # スタイル適用
                self.apply_excel_styles(writer, output_path)
            
            self.status_var.set(f"保存完了: {output_path}")
            messagebox.showinfo("完了", f"検索結果を保存しました。\n{len(self.search_results)}件のデータを出力\n\n保存先: {output_path}")
            
        except Exception as e:
            messagebox.showerror("エラー", f"保存エラー: {e}")
            self.status_var.set("保存エラー")
    
    def apply_excel_styles(self, writer, output_path):
        """Excelファイルにスタイルを適用"""
        try:
            # ワークブックを取得
            workbook = writer.book
            
            # 検索結果シートのスタイル設定
            if '検索結果' in workbook.sheetnames:
                worksheet = workbook['検索結果']
                
                # ヘッダー行のスタイル
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                
                for cell in worksheet[1]:  # 1行目（ヘッダー）
                    cell.font = header_font
                    cell.fill = header_fill
                
                # 列幅の自動調整
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
        except Exception as e:
            print(f"スタイル適用エラー: {e}")
    
    def clear_all(self):
        """全てクリア"""
        self.df = None
        self.search_results = None
        self.columns = []
        
        self.input_file_path.set("")
        self.output_file_path.set("")
        self.search_keyword.set("")
        self.column_combo['values'] = []
        self.search_column.set("")
        
        self.data_text.delete(1.0, tk.END)
        self.result_text.delete(1.0, tk.END)
        
        self.status_var.set("クリアしました")
    
    def create_sample(self):
        """サンプルExcelファイルを作成"""
        try:
            # サンプルデータ作成
            sample_data = {
                '会社名': [
                    '株式会社サンプル',
                    '有限会社テスト商事',
                    'サンプル工業株式会社',
                    '株式会社テストシステム',
                    'デモ株式会社',
                    '株式会社サンプルテクノロジー',
                    'テスト商事有限会社',
                    'サンプル・デザイン株式会社',
                    '株式会社テストソリューション',
                    'サンプル物産株式会社',
                    'Apple Japan株式会社',
                    'アップル販売株式会社',
                    'マイクロソフト株式会社',
                    'Microsoft Japan',
                    'Google Japan株式会社'
                ],
                '代表者': [
                    '田中太郎',
                    '佐藤花子',
                    '山田次郎',
                    '鈴木美咲',
                    '高橋健一',
                    '伊藤愛子',
                    '渡辺直樹',
                    '中村由美',
                    '小林修一',
                    '加藤雅子',
                    'Smith John',
                    'Johnson Mary',
                    'Williams David',
                    'Brown Lisa',
                    'Jones Michael'
                ],
                '業種': [
                    'IT・ソフトウェア',
                    '商社・貿易',
                    '製造業',
                    'IT・システム開発',
                    'コンサルティング',
                    'IT・AI開発',
                    '商社・卸売',
                    'デザイン・広告',
                    'IT・ソリューション',
                    '商社・物流',
                    'IT・ハードウェア',
                    'IT・販売',
                    'IT・ソフトウェア',
                    'IT・クラウド',
                    'IT・検索エンジン'
                ],
                '従業員数': [150, 25, 300, 80, 12, 200, 30, 45, 120, 180, 500, 100, 800, 1000, 2000],
                '所在地': [
                    '東京都渋谷区',
                    '大阪府大阪市',
                    '愛知県名古屋市',
                    '東京都新宿区',
                    '福岡県福岡市',
                    '東京都港区',
                    '大阪府堺市',
                    '東京都品川区',
                    '神奈川県横浜市',
                    '東京都千代田区',
                    '東京都港区',
                    '東京都中央区',
                    '東京都港区',
                    '東京都渋谷区',
                    '東京都渋谷区'
                ]
            }
            
            # DataFrameに変換
            df = pd.DataFrame(sample_data)
            
            # サンプルファイル保存
            sample_file = "sample_companies.xlsx"
            df.to_excel(sample_file, index=False, engine='openpyxl')
            
            # 自動的にファイルパスを設定
            self.input_file_path.set(sample_file)
            self.output_file_path.set("sample_search_results.xlsx")
            
            # ファイルを読み込み
            self.load_excel_file()
            
            self.status_var.set("サンプルファイルを作成しました")
            messagebox.showinfo("完了", f"サンプルファイルを作成しました: {sample_file}\n\n🔍 試してみよう:\n• 完全一致: 'Apple Japan株式会社'\n• 部分一致: 'サンプル'\n• 曖昧検索: 'あっぷる'（類似度70%）")
            
        except Exception as e:
            messagebox.showerror("エラー", f"サンプル作成エラー: {e}")

def main():
    """メイン関数"""
    root = tk.Tk()
    app = ExcelSearchTool(root)
    
    # スタイル設定
    style = ttk.Style()
    try:
        style.configure("Accent.TButton", foreground="white", background="blue")
    except:
        pass  # スタイル設定に失敗しても続行
    
    root.mainloop()

if __name__ == '__main__':
    main()
